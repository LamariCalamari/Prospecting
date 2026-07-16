"""Unit tests for the framework-free core logic (no network needed).

Run from the project root:  python -m unittest discover -s tests
"""
import unittest

from core import assembly, export, prospecting, valuation
from core.models import (
    Appointment,
    CompanyInfo,
    NewsItem,
    OfficerCandidate,
    OneSheet,
    PSC,
    WikidataFacts,
    WikiSummary,
)
from sources import companies_house as ch


class TestCompaniesHouseHelpers(unittest.TestCase):
    def test_officer_id_from_self_link(self):
        self.assertEqual(
            ch._officer_id_from_self_link("/officers/AbC123/appointments"), "AbC123"
        )
        self.assertEqual(ch._officer_id_from_self_link("/officers/XyZ"), "XyZ")

    def test_format_dob_month_and_year(self):
        self.assertEqual(ch._format_dob({"year": 1980, "month": 6}), "1980-06")
        self.assertEqual(ch._format_dob({"year": 1980}), "1980")
        self.assertIsNone(ch._format_dob(None))

    def test_address_to_str_skips_blanks(self):
        addr = {"address_line_1": "1 Road", "locality": "London", "region": None}
        self.assertEqual(ch._address_to_str(addr), "1 Road, London")
        self.assertIsNone(ch._address_to_str({}))


class TestPaging(unittest.TestCase):
    def test_paged_items_stops_at_total(self):
        # Two pages of 2, total 3 -> should stop after collecting 3.
        pages = [
            {"items": [{"i": 1}, {"i": 2}], "total_results": 3},
            {"items": [{"i": 3}], "total_results": 3},
        ]
        calls = {"n": 0}

        def fake_get(path, params=None):
            page = pages[calls["n"]]
            calls["n"] += 1
            return page

        original = ch._get
        ch._get = fake_get
        try:
            items = ch._paged_items("/x", max_items=100, page_size=2)
        finally:
            ch._get = original
        self.assertEqual([d["i"] for d in items], [1, 2, 3])

    def test_paged_items_respects_max(self):
        def fake_get(path, params=None):
            return {"items": [{"i": 1}, {"i": 2}], "total_results": 999}

        original = ch._get
        ch._get = fake_get
        try:
            items = ch._paged_items("/x", max_items=2, page_size=2)
        finally:
            ch._get = original
        self.assertEqual(len(items), 2)


class TestNewsQuery(unittest.TestCase):
    def test_prefers_company_signal(self):
        officer = OfficerCandidate(
            officer_id="x", name="Jane", source_url="", top_companies=["ACME LTD"]
        )
        q = assembly._news_query("Jane Example", officer, "some context")
        self.assertIn('"Jane Example"', q)
        self.assertIn("ACME LTD", q)

    def test_falls_back_to_context(self):
        q = assembly._news_query("Jane Example", None, "London fintech")
        self.assertIn("London fintech", q)


class TestExport(unittest.TestCase):
    def _sheet(self):
        return OneSheet(
            confirmed_name="Jane Example",
            context="CTO",
            appointments=[
                Appointment("ACME LTD", "12345678", "director", "active", "2020-01-01",
                            None, None, "http://ch/company/12345678")
            ],
            psc_filings=[
                PSC("ACME LTD", "12345678", "Jane Example", "individual",
                    ["ownership-of-shares-75-to-100-percent"], "2020-01-01", None, "http://ch")
            ],
            companies=[CompanyInfo("ACME LTD", "12345678", "active", source_url="http://ch")],
            wiki=WikiSummary("Jane", "She said “hello”—loudly.", "entrepreneur", "http://w"),
            news=[NewsItem("Big news", "http://n", "The Times", "2026")],
        )

    def test_markdown_attributes_and_no_invention(self):
        md = export.to_markdown(self._sheet())
        self.assertIn("Companies House", md)
        self.assertIn("ownership-of-shares-75-to-100-percent", md)  # verbatim, not computed
        self.assertIn("not verified claims", md)  # news disclaimer present

    def test_markdown_blank_when_no_wiki(self):
        sheet = OneSheet(confirmed_name="Nobody")
        md = export.to_markdown(sheet)
        self.assertIn("No confident Wikipedia match", md)
        self.assertIn("No directorships returned", md)

    def test_pdf_is_valid_and_handles_unicode(self):
        pdf = export.to_pdf(self._sheet())
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 800)

    def test_pdf_safe_normalizes_smart_punctuation(self):
        out = export._pdf_safe("She said “hi” — don’t…")
        self.assertNotIn("?", out)
        self.assertIn('"hi"', out)
        self.assertIn("don't", out)

    def test_excel_is_valid_xlsx_with_expected_tabs(self):
        from io import BytesIO

        from openpyxl import load_workbook

        data = export.to_excel(self._sheet())
        self.assertTrue(data[:2] == b"PK")  # xlsx is a zip
        wb = load_workbook(BytesIO(data))
        for tab in ["Summary", "Directorships", "Ownership stakes", "Companies",
                    "PSC (all)", "News"]:
            self.assertIn(tab, wb.sheetnames)
        # Directorship data landed in the right tab.
        dvals = [c.value for row in wb["Directorships"].iter_rows() for c in row]
        self.assertIn("ACME LTD", dvals)


class TestProspectingSignals(unittest.TestCase):
    def test_humanize_control_band(self):
        self.assertEqual(
            prospecting.humanize_control("ownership-of-shares-75-to-100-percent"),
            "Ownership of shares 75–100%",
        )
        self.assertEqual(
            prospecting.humanize_control("significant-influence-or-control"),
            "Significant influence or control",
        )

    def test_names_match_handles_order_and_titles(self):
        self.assertTrue(prospecting.names_match("STORONSKY Nikolay", "Nikolay Storonsky"))
        self.assertTrue(prospecting.names_match("Mr Nikolay Storonsky", "Nikolay Storonsky"))
        self.assertFalse(prospecting.names_match("Jane Smith", "John Doe"))

    def test_derive_signals_picks_top_stake_and_matches_person(self):
        sheet = OneSheet(
            confirmed_name="Jane Example",
            appointments=[
                Appointment("ACME LTD", "1", "director", "active"),
                Appointment("OLD LTD", "2", "director", "resigned"),
            ],
            psc_filings=[
                PSC("ACME LTD", "1", "Jane Example", "individual",
                    ["ownership-of-shares-25-to-50-percent"]),
                PSC("BIGCO LTD", "3", "Jane Example", "individual",
                    ["ownership-of-shares-75-to-100-percent"]),
                PSC("OTHER LTD", "4", "Someone Else", "individual",
                    ["ownership-of-shares-75-to-100-percent"]),
            ],
            companies=[CompanyInfo("ACME LTD", "1", has_charges=True)],
            wikidata=WikidataFacts(qid="Q1", source_url="http://wd",
                                   net_worth="1,000,000 pound", linkedin="http://li/jane"),
        )
        sig = prospecting.derive_signals(sheet, "Jane Example")
        self.assertEqual(sig.active_directorships, 1)
        self.assertEqual(sig.resigned_directorships, 1)
        self.assertEqual(len(sig.stakes), 2)  # excludes "Someone Else"
        self.assertEqual(sig.top_ownership, "Ownership of shares 75–100%")
        self.assertEqual(sig.companies_with_charges, 1)
        self.assertEqual(sig.net_worth, "1,000,000 pound")
        self.assertEqual(sig.linkedin_url, "http://li/jane")
        self.assertIn("linkedin.com", sig.linkedin_search_url)


class TestAutoSelect(unittest.TestCase):
    def _off(self, name, dob=None, appts=None):
        return OfficerCandidate(officer_id="x", name=name, source_url="",
                                date_of_birth=dob, appointment_count=appts)

    def test_best_officer_picks_matching_birth_year(self):
        officers = [
            self._off("CANDY CREATIONS UK LTD", None, 1),          # corporate — skip
            self._off("Nicholas Anthony Christopher CANDY", "1955-06", 7),  # namesake
            self._off("Nicholas Anthony Christopher CANDY", "1973-01", 3),  # the one
        ]
        idx = prospecting.best_officer_index(officers, "Nick Candy", birth_year="1973")
        self.assertEqual(idx, 2)  # nickname + birth year pin the right Candy

    def test_best_officer_none_when_ambiguous_without_year(self):
        officers = [
            self._off("Nicholas CANDY", "1955-06", 7),
            self._off("Nicholas CANDY", "1973-01", 3),
        ]
        # Two same-name people, no birth year to disambiguate -> don't guess.
        self.assertIsNone(prospecting.best_officer_index(officers, "Nicholas Candy"))

    def test_best_officer_none_when_no_name_match(self):
        officers = [self._off("John Smith", "1960-01", 3)]
        self.assertIsNone(prospecting.best_officer_index(officers, "Jane Doe"))

    def test_extract_birth_year(self):
        self.assertEqual(
            prospecting.extract_birth_year("British property developer (born 1973)"),
            "1973",
        )
        self.assertIsNone(prospecting.extract_birth_year("no year here"))

    def test_nickname_name_match(self):
        self.assertTrue(prospecting.names_match("Nick Candy", "Nicholas Candy"))
        self.assertTrue(prospecting.names_match("Nicholas Anthony CANDY", "Nick Candy"))

    def test_best_wiki_index_matches_title(self):
        class W:
            def __init__(self, t):
                self.title = t
        cands = [W("Candy (disambiguation)"), W("Nick Candy")]
        self.assertEqual(prospecting.best_wiki_index(cands, "Nick Candy"), 1)


class TestDisplayHelpers(unittest.TestCase):
    def test_company_case_keeps_short_acronyms(self):
        self.assertEqual(
            prospecting.company_case("CANDY & CANDY HOLDINGS LIMITED"),
            "Candy & Candy Holdings Limited",
        )
        self.assertEqual(
            prospecting.company_case("NC (LONDON) LIMITED"), "NC (London) Limited"
        )
        self.assertEqual(prospecting.company_case("49 UBS LIMITED"), "49 UBS Limited")

    def test_short_band(self):
        self.assertEqual(
            prospecting.short_band("Ownership of shares 75–100%"), "75–100% shares"
        )
        self.assertIsNone(prospecting.short_band(None))


class TestOfficerGrouping(unittest.TestCase):
    def _off(self, name, dob=None, appts=None, oid="x"):
        return OfficerCandidate(officer_id=oid, name=name, source_url="",
                                date_of_birth=dob, appointment_count=appts)

    def test_fragments_collapse_into_one_group(self):
        officers = [
            self._off("Nicholas Anthony Christopher CANDY", "1973-01", 1, "a"),
            self._off("Nicholas Anthony Christopher CANDY", None, 1, "b"),
            self._off("Nicholas Anthony Christopher CANDY", "1973-01", 45, "c"),
            self._off("Nicholas Anthony Christopher CANDY", "1955-06", 7, "d"),
        ]
        groups = prospecting.group_officers(officers)
        self.assertEqual(len(groups), 2)  # 1973 person (+dob-less) and 1955 person
        main = groups[0]
        self.assertEqual(main.birth_year, "1973")
        self.assertEqual(len(main.records), 3)
        self.assertEqual(main.primary.officer_id, "c")  # fattest record represents

    def test_best_group_index_uses_birth_year(self):
        officers = [
            self._off("Nicholas CANDY", "1955-06", 7, "a"),
            self._off("Nicholas CANDY", "1973-01", 1, "b"),
        ]
        groups = prospecting.group_officers(officers)
        idx = prospecting.best_group_index(groups, "Nick Candy", "1973")
        self.assertEqual(groups[idx].birth_year, "1973")
        # Ambiguous without a year:
        self.assertIsNone(prospecting.best_group_index(groups, "Nick Candy"))


class TestMergePeople(unittest.TestCase):
    def _grp(self, name, year, appts, oid="x"):
        g = prospecting.group_officers([
            OfficerCandidate(officer_id=oid, name=name, source_url="",
                             date_of_birth=f"{year}-01" if year else None,
                             appointment_count=appts)
        ])[0]
        return g

    def test_wiki_person_gets_matching_ch_group(self):
        wiki_persons = [{
            "title": "Nick Candy",
            "description": "British luxury property developer",
            "thumbnail": None,
            "birth_year": "1973",
            "lead_name": "Nicholas Anthony Christopher Candy",
        }]
        groups = [
            self._grp("Nicholas Anthony Christopher CANDY", "1955", 7, "old"),
            self._grp("Nicholas Anthony Christopher CANDY", "1973", 45, "him"),
        ]
        people = assembly.merge_people(wiki_persons, groups, "Nick Candy")
        self.assertEqual(people[0].display_name, "Nick Candy")
        self.assertTrue(people[0].has_wiki and people[0].has_ch)
        self.assertEqual(people[0].officer.officer_id, "him")  # 1973, not 1955
        # The 1955 namesake still appears as a CH-only candidate below.
        ch_only = [p for p in people[1:] if p.has_ch and not p.has_wiki]
        self.assertTrue(any(p.birth_year == "1955" for p in ch_only))

    def test_wiki_only_person_still_listed(self):
        wiki_persons = [{"title": "Jane Doe", "description": "novelist",
                         "thumbnail": None, "birth_year": None, "lead_name": None}]
        people = assembly.merge_people(wiki_persons, [], "Jane Doe")
        self.assertEqual(len(people), 1)
        self.assertTrue(people[0].has_wiki)
        self.assertFalse(people[0].has_ch)


class TestValuation(unittest.TestCase):
    def test_band_to_range(self):
        self.assertEqual(
            valuation.band_to_range(["ownership-of-shares-75-to-100-percent"]),
            (0.75, 1.0),
        )
        self.assertEqual(
            valuation.band_to_range(
                ["voting-rights-25-to-50-percent",
                 "ownership-of-shares-25-to-50-percent"]
            ),
            (0.25, 0.5),
        )
        self.assertEqual(
            valuation.band_to_range(
                ["ownership-of-shares-more-than-25-percent-registered-overseas-entity"]
            ),
            (0.25, 1.0),
        )
        self.assertIsNone(valuation.band_to_range(["significant-influence-or-control"]))

    def test_extract_ixbrl_figures(self):
        xhtml = (
            '<html><body>'
            '<ix:nonFraction name="uk-core:NetAssetsLiabilities" contextRef="c1" '
            'unitRef="GBP" decimals="0" scale="3" format="ixt:numdotdecimal">'
            '1,234</ix:nonFraction>'
            '<ix:nonFraction name="uk-core:CashBankOnHand" contextRef="c1" '
            'unitRef="GBP" sign="-" decimals="0">500</ix:nonFraction>'
            '</body></html>'
        )
        fig = valuation.extract_ixbrl_figures(xhtml)
        self.assertEqual(fig["net_assets"], 1_234_000)  # scale=3
        self.assertEqual(fig["cash"], -500)             # sign=-

    def test_build_estimates_math_and_clamp(self):
        sheet = OneSheet(
            confirmed_name="Jane Example",
            psc_filings=[
                PSC("GOODCO", "1", "Jane Example", "individual",
                    ["ownership-of-shares-50-to-75-percent"]),
                PSC("BADCO", "2", "Jane Example", "individual",
                    ["ownership-of-shares-75-to-100-percent"]),
            ],
            companies=[
                CompanyInfo("GOODCO", "1", net_assets=1_000_000.0,
                            accounts_last_made_up_to="2025-12-31"),
                CompanyInfo("BADCO", "2", net_assets=-50_000.0),  # net liabilities
            ],
        )
        est = valuation.build_estimates(sheet, "Jane Example")
        self.assertEqual(est.counted, 2)
        good = next(s for s in est.stakes if s.company_name == "GOODCO")
        self.assertEqual((good.value_lo, good.value_hi), (500_000.0, 750_000.0))
        bad = next(s for s in est.stakes if s.company_name == "BADCO")
        self.assertEqual((bad.value_lo, bad.value_hi), (0.0, 0.0))  # clamped
        self.assertEqual(est.total_hi, 750_000.0)

    def test_fmt_gbp(self):
        self.assertEqual(valuation.fmt_gbp(1_500_000_000), "£1.5bn")
        self.assertEqual(valuation.fmt_gbp(2_300_000), "£2.3m")
        self.assertEqual(valuation.fmt_gbp(45_000), "£45k")
        self.assertEqual(valuation.fmt_gbp(None), "—")


class TestNewSignalSources(unittest.TestCase):
    def test_ec_date_parse(self):
        from sources.electoral_commission import parse_ec_date
        self.assertEqual(parse_ec_date("/Date(1759276800000)/"), "2025-10-01")
        self.assertIsNone(parse_ec_date(None))

    def test_market_same_company(self):
        from sources.market import same_company
        self.assertTrue(same_company("CERILLION TECHNOLOGIES LIMITED", "Cerillion Plc"))
        self.assertTrue(same_company("CERILLION PLC", "Cerillion Plc"))
        self.assertFalse(same_company("CANDY CAPITAL LIMITED", "Cerillion Plc"))

    def test_ixbrl_extracts_pnl_concepts(self):
        xhtml = (
            '<ix:nonFraction name="uk:TurnoverRevenue" contextRef="c" scale="3">'
            "43,800</ix:nonFraction>"
            '<ix:nonFraction name="uk:ProfitLossOnOrdinaryActivitiesBeforeTax" '
            'contextRef="c" scale="3">12,300</ix:nonFraction>'
            '<ix:nonFraction name="uk:AverageNumberEmployeesDuringPeriod" '
            'contextRef="c">305</ix:nonFraction>'
        )
        fig = valuation.extract_ixbrl_figures(xhtml)
        self.assertEqual(fig["turnover"], 43_800_000)
        self.assertEqual(fig["profit"], 12_300_000)
        self.assertEqual(fig["employees"], 305)

    def test_why_they_matter_bullets(self):
        from core import signals
        from core.models import Donation, ListedQuote
        sheet = OneSheet(
            confirmed_name="Jane Example",
            listed=[ListedQuote("EXAMPLE PLC", "EXM.L", "LSE", 500.0, "GBp",
                                "http://y/EXM.L")],
            donations=[Donation("Jane Example", "Some Party", 100_000.0,
                                "2025-01-01", "http://ec")],
            companies=[CompanyInfo("EXAMPLE PLC", "1", turnover=43_800_000.0,
                                   employees=305, source_url="http://ch/1")],
        )
        sig = prospecting.derive_signals(sheet, "Jane Example")
        est = valuation.build_estimates(sheet, "Jane Example")
        bullets = signals.why_they_matter(sheet, sig, est)
        text = "\n".join(bullets)
        self.assertIn("publicly listed", text)
        self.assertIn("£43.8m", text)
        self.assertIn("£100k", text)
        self.assertIn("Some Party", text)


if __name__ == "__main__":
    unittest.main()
