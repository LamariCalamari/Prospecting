"""Export a OneSheet to Markdown or PDF.

Markdown is the source of truth; the PDF is a plain rendering of the same
content using fpdf2 (pure-Python, no system dependencies). Both attribute
every fact with a source link, matching the on-screen sheet.
"""
from __future__ import annotations

from core import prospecting, signals, valuation
from core.models import OneSheet

# fpdf2's core fonts are Latin-1 only. Map the common "smart" punctuation that
# shows up in Wikipedia/news text to plain equivalents so the PDF renders them
# instead of "?" placeholders.
_UNICODE_FIXES = {
    "‘": "'", "’": "'", "“": '"', "”": '"',
    "–": "-", "—": "-", "…": "...", " ": " ",
    "•": "-", "′": "'", "″": '"',
}


def _pdf_safe(text: str) -> str:
    for bad, good in _UNICODE_FIXES.items():
        text = text.replace(bad, good)
    return text.encode("latin-1", "replace").decode("latin-1")


def _fmt_date(value) -> str:
    return value or ""


def to_markdown(sheet: OneSheet) -> str:
    lines: list[str] = []
    lines.append(f"# Prospecting one-sheet — {sheet.confirmed_name}")
    if sheet.context:
        lines.append(f"*Research context:* {sheet.context}")
    lines.append("")
    lines.append(
        "> Every fact below is attributed to its source. Blank fields mean the "
        "source returned nothing — nothing here is inferred or estimated."
    )
    lines.append("")

    # Prospecting snapshot
    sig = prospecting.derive_signals(sheet, sheet.confirmed_name)
    lines.append("## Prospecting snapshot")
    lines.append(f"- **Net worth (published):** {sig.net_worth or '—'}")
    stake = sig.top_ownership or (
        f"{sig.top_former_ownership} (former)" if sig.top_former_ownership else "—"
    )
    lines.append(f"- **Largest disclosed stake:** {stake}")
    lines.append(f"- **Active directorships:** {sig.active_directorships}")
    lines.append(f"- **Companies controlled (PSC):** {len(sig.stakes)} current, "
                 f"{len(sig.former_stakes)} former")
    if sig.companies_with_charges:
        lines.append(f"- **Companies with registered charges (debt):** {sig.companies_with_charges}")
    if sig.insolvency_companies:
        lines.append(f"- **Companies with insolvency history:** {sig.insolvency_companies}")
    links = []
    if sig.linkedin_url:
        links.append(f"[LinkedIn (verified)]({sig.linkedin_url})")
    links.append(f"[Search LinkedIn]({sig.linkedin_search_url})")
    links.append(f"[Google → LinkedIn]({sig.google_linkedin_url})")
    lines.append("- **Find online:** " + " · ".join(links))
    lines.append(
        "_Net worth shows only when published. Stake = Companies House PSC band, "
        "verbatim._"
    )
    lines.append("")

    # Why they matter — synthesized, each bullet cites its source.
    est_full = valuation.build_estimates(sheet, sheet.confirmed_name)
    why = signals.why_they_matter(sheet, sig, est_full)
    if why:
        lines.append("## Why they matter")
        for bullet in why:
            lines.append(f"- {bullet}")
        lines.append("")

    # Political donations
    if sheet.donations:
        lines.append("## Political donations — source: Electoral Commission")
        for d in sheet.donations:
            val = valuation.fmt_gbp(d.value) if d.value else "—"
            lines.append(
                f"- {val} to {d.recipient} ({d.date or '—'}) — donor on record: "
                f"{d.donor_name} — [EC register]({d.source_url})"
            )
        lines.append("")

    # Estimated stake value (book-value basis)
    est = valuation.build_estimates(sheet, sheet.confirmed_name)
    if est.stakes:
        lines.append("## Estimated stake value (book-value basis)")
        lines.append(
            "_Method: disclosed stake band (CH PSC) × net assets from the latest "
            "filed accounts. A conservative floor, not market value. Estimates, "
            "not facts._"
        )
        if est.counted:
            lines.append(
                f"- **Total: {valuation.fmt_gbp(est.total_lo)} – "
                f"{valuation.fmt_gbp(est.total_hi)}** "
                f"(across {est.counted} of {len(est.stakes)} stakes with "
                "machine-readable accounts)"
            )
        for s in est.stakes:
            band = f"{s.stake_lo:.0%}–{s.stake_hi:.0%}"
            if s.net_assets is not None:
                lines.append(
                    f"- {s.company_name}: {band} × net assets "
                    f"{valuation.fmt_gbp(s.net_assets)}"
                    + (f" (accounts to {s.accounts_date})" if s.accounts_date else "")
                    + f" = {valuation.fmt_gbp(s.value_lo)}–{valuation.fmt_gbp(s.value_hi)}"
                    f" — [Companies House]({s.source_url})"
                )
            else:
                lines.append(
                    f"- {s.company_name}: {band} × (accounts not machine-readable)"
                    f" — [Companies House]({s.source_url})"
                )
        lines.append("")

    # Position & company
    lines.append("## Position & company")
    active = [a for a in sheet.appointments if a.status == "active"]
    if active:
        for a in active[:5]:
            role = a.officer_role or "officer"
            lines.append(
                f"- **{role.title()}**, {a.company_name} "
                f"(appointed {_fmt_date(a.appointed_on)}) "
                f"— [Companies House]({a.source_url})"
            )
    if sheet.wiki and sheet.wiki.description:
        lines.append(f"- {sheet.wiki.description} — [Wikipedia]({sheet.wiki.url})")
    if not active and not (sheet.wiki and sheet.wiki.description):
        lines.append("_No current position data returned._")
    lines.append("")

    # Directorships & control
    lines.append("## Directorships & control")
    if sheet.appointments:
        lines.append("### Appointments")
        for a in sheet.appointments:
            status = a.status or ""
            dates = _fmt_date(a.appointed_on)
            if a.resigned_on:
                dates += f" → {a.resigned_on}"
            role = a.officer_role or "officer"
            lines.append(
                f"- {a.company_name} — {role} ({status}); {dates} "
                f"— [Companies House]({a.source_url})"
            )
    else:
        lines.append("_No directorships returned._")

    if sheet.psc_filings:
        lines.append("")
        lines.append("### Persons with significant control (PSC)")
        for p in sheet.psc_filings:
            controls = "; ".join(p.natures_of_control) if p.natures_of_control else "—"
            ceased = f", ceased {p.ceased_on}" if p.ceased_on else ""
            lines.append(
                f"- {p.name or '(unnamed)'} @ {p.company_name}: {controls} "
                f"(notified {_fmt_date(p.notified_on)}{ceased}) "
                f"— [Companies House]({p.source_url})"
            )
    lines.append("")

    # Background
    lines.append("## Background")
    if sheet.wiki:
        lines.append(sheet.wiki.extract)
        lines.append("")
        lines.append(f"— [Wikipedia: {sheet.wiki.title}]({sheet.wiki.url})")
    else:
        lines.append("_No confident Wikipedia match._")
    lines.append("")

    # Key facts (Wikidata)
    wd = sheet.wikidata
    if wd:
        lines.append("### Key facts — source: Wikidata")
        if wd.net_worth:
            lines.append(f"- Net worth: {wd.net_worth}")
        if wd.occupations:
            lines.append(f"- Occupation: {', '.join(wd.occupations)}")
        if wd.positions:
            lines.append(f"- Positions held: {', '.join(wd.positions)}")
        if wd.educated_at:
            lines.append(f"- Educated at: {', '.join(wd.educated_at)}")
        if wd.official_website:
            lines.append(f"- Official website: {wd.official_website}")
        if wd.linkedin:
            lines.append(f"- LinkedIn: {wd.linkedin}")
        if wd.twitter:
            lines.append(f"- X/Twitter: {wd.twitter}")
        lines.append(f"— [Wikidata {wd.qid}]({wd.source_url})")
        lines.append("")

    # Funding / company info
    lines.append("## Funding / company info")
    if wd and wd.official_website:
        lines.append(f"- Official website: {wd.official_website} (source: Wikidata)")
    if sheet.companies:
        for c in sheet.companies:
            bits = [f"**{c.company_name}** ({c.company_number})"]
            if c.status:
                bits.append(f"status: {c.status}")
            if c.incorporation_date:
                bits.append(f"incorporated: {c.incorporation_date}")
            lines.append("- " + ", ".join(bits) + f" — [Companies House]({c.source_url})")
            sig = []
            if c.accounts_last_made_up_to:
                sig.append(f"accounts to {c.accounts_last_made_up_to}")
            if c.accounts_next_due:
                sig.append(f"next due {c.accounts_next_due}" + (" (overdue)" if c.accounts_overdue else ""))
            if c.has_charges:
                sig.append(f"{len(c.charges) or 'has'} charge(s)")
            if c.has_insolvency_history:
                sig.append("insolvency history")
            if sig:
                lines.append(f"  - signals: {' · '.join(sig)}")
    else:
        lines.append("_No company profiles returned._")
    lines.append("")

    # Regulatory record (FCA)
    lines.append("## Regulatory record (FCA)")
    if sheet.fca_records:
        for r in sheet.fca_records:
            status = f" — {r.status}" if r.status else ""
            lines.append(f"- {r.name} (IRN {r.reference_number or '—'}){status} — [FCA]({r.source_url})")
            if r.roles:
                lines.append(f"  - roles: {', '.join(r.roles)}")
            if r.firms:
                lines.append(f"  - firms: {', '.join(r.firms)}")
    else:
        lines.append("_No FCA-approved individual matched (or not configured)._")
    lines.append("")

    # Philanthropy (Charity Commission)
    lines.append("## Philanthropy (Charity Commission)")
    if sheet.charities:
        for ch_rec in sheet.charities:
            num = f" (no. {ch_rec.charity_number})" if ch_rec.charity_number else ""
            lines.append(f"- {ch_rec.name}{num} — [Charity register]({ch_rec.source_url})")
    else:
        lines.append("_No matching charities (or not configured)._")
    lines.append("")

    # Risk & due diligence
    lines.append("## Risk & due diligence")
    lines.append("_Screening leads to verify — a name match is not a confirmed identification._")
    if sheet.sanctions_hits:
        lines.append("### Sanctions / PEP — source: OpenSanctions")
        for h in sheet.sanctions_hits:
            topics = f" — {', '.join(h.topics)}" if h.topics else ""
            lines.append(f"- {h.name} ({h.schema or '—'}){topics} — [OpenSanctions]({h.source_url})")
    if sheet.gazette_notices:
        lines.append("### Official notices — source: The Gazette")
        for g in sheet.gazette_notices:
            pub = f" — {g.published[:10]}" if g.published else ""
            lines.append(f"- [{g.title}]({g.link}){pub}")
    if not sheet.sanctions_hits and not sheet.gazette_notices:
        lines.append("_No screening matches._")
    lines.append("")

    # News & controversies
    lines.append("## News & controversies")
    lines.append(
        "_Headlines surfaced from a news search for review — not verified claims._"
    )
    if sheet.news:
        for n in sheet.news:
            src = f" — {n.source}" if n.source else ""
            lines.append(f"- [{n.title}]({n.link}){src}")
    else:
        lines.append("_No news results._")
    lines.append("")

    return "\n".join(lines)


def to_pdf(sheet: OneSheet) -> bytes:
    """Render the markdown content to a simple, clean PDF via fpdf2."""
    from fpdf import FPDF

    md = to_markdown(sheet)

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(15, 15, 15)
    effective_width = pdf.w - 30

    def write_line(text: str, size: int, style: str = "", gap: float = 2):
        pdf.set_font("Helvetica", style, size)
        pdf.multi_cell(effective_width, size * 0.5 + 2, _pdf_safe(text))
        pdf.ln(gap)

    for raw in md.splitlines():
        line = raw.rstrip()
        if not line:
            pdf.ln(2)
        elif line.startswith("# "):
            write_line(line[2:], 18, "B", gap=3)
        elif line.startswith("## "):
            write_line(line[3:], 14, "B", gap=2)
        elif line.startswith("### "):
            write_line(line[4:], 12, "B", gap=1)
        elif line.startswith(("> ", "_")):
            write_line(line.lstrip("> "), 9, "I")
        else:
            write_line(line, 10)

    out = pdf.output()  # fpdf2 returns a bytearray
    return bytes(out)


def to_excel(sheet: OneSheet) -> bytes:
    """Render the one-sheet as a structured multi-tab .xlsx workbook.

    One tab per data domain (Summary, Directorships, Ownership, Companies, PSC,
    Regulatory, Charities, Sanctions, Gazette, News). Every tab has a styled,
    frozen header row, auto-sized columns, and clickable source links — so the
    export is analysis-ready, not just a dump. Empty domains still get a tab with
    a clear "no data" note, so the structure is predictable.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    link_font = Font(color="2563EB", underline="single")
    wrap = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()

    def _add_sheet(name: str, headers: list[str], rows: list[list],
                   link_cols: Optional[set[int]] = None, note: str = ""):
        ws = wb.create_sheet(title=name[:31])
        link_cols = link_cols or set()
        if not rows:
            ws["A1"] = note or "No data returned for this source."
            ws["A1"].font = Font(italic=True, color="6B7280")
            return ws
        for c, head in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=head)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c)
                text = "" if val is None else str(val)
                cell.value = text
                cell.alignment = wrap
                if c in link_cols and text.startswith("http"):
                    cell.hyperlink = text
                    cell.font = link_font
        # Auto width (capped) from the longest cell in each column.
        for c in range(1, len(headers) + 1):
            longest = max(
                [len(str(headers[c - 1]))]
                + [len(str(row[c - 1])) for row in rows if c - 1 < len(row)]
            )
            ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 12), 60)
        ws.freeze_panes = "A2"
        return ws

    sig = prospecting.derive_signals(sheet, sheet.confirmed_name)

    # --- Summary tab (key/value) ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Prospecting one-sheet — {sheet.confirmed_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    stake = sig.top_ownership or (
        f"{sig.top_former_ownership} (former)" if sig.top_former_ownership else "—"
    )
    est_sum = valuation.build_estimates(sheet, sheet.confirmed_name)
    book = (
        f"{valuation.fmt_gbp(est_sum.total_lo)} – {valuation.fmt_gbp(est_sum.total_hi)}"
        if est_sum.counted else "—"
    )
    donation_total = sum(d.value for d in sheet.donations if d.value)
    listed_desc = "; ".join(
        f"{q.company_name} ({q.symbol})" for q in sheet.listed
    ) or "—"
    summary_rows = [
        ("Name", sheet.confirmed_name),
        ("Research context", sheet.context or "—"),
        ("Net worth (published)", sig.net_worth or "—"),
        ("Largest disclosed stake", stake),
        ("Est. stake value (book-value floor)", book),
        ("Political donations (registered)",
         f"£{donation_total:,.0f}" if donation_total else "—"),
        ("Listed-company roles", listed_desc),
        ("Active directorships", sig.active_directorships),
        ("Past directorships", sig.resigned_directorships),
        ("Companies controlled (current PSC)", len(sig.stakes)),
        ("Companies controlled (former PSC)", len(sig.former_stakes)),
        ("Companies with charges (debt)", sig.companies_with_charges),
        ("Companies with insolvency history", sig.insolvency_companies),
        ("Occupation", ", ".join(sheet.wikidata.occupations) if sheet.wikidata else "—"),
        ("LinkedIn (verified)", sig.linkedin_url or "—"),
        ("Search LinkedIn", sig.linkedin_search_url),
        ("Official website",
         sheet.wikidata.official_website if sheet.wikidata and sheet.wikidata.official_website else "—"),
        ("Wikipedia", sheet.wiki.url if sheet.wiki else "—"),
    ]
    for r, (k, v) in enumerate(summary_rows, start=3):
        kc = ws.cell(row=r, column=1, value=k)
        kc.font = Font(bold=True)
        vcell = ws.cell(row=r, column=2, value=str(v))
        vcell.alignment = wrap
        if str(v).startswith("http"):
            vcell.hyperlink = str(v)
            vcell.font = link_font
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 70
    ws["A2"] = ("Every value is sourced. Blanks/'—' mean the source returned "
                "nothing — nothing is inferred or estimated.")
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:B2")

    # --- Directorships ---
    _add_sheet(
        "Directorships",
        ["Status", "Company", "Company number", "Role", "Appointed", "Resigned", "Source"],
        [[a.status, a.company_name, a.company_number, a.officer_role,
          a.appointed_on, a.resigned_on, a.source_url] for a in sheet.appointments],
        link_cols={7},
        note="No directorships returned (no Companies House officer selected, or none on file).",
    )

    # --- Ownership stakes ---
    _add_sheet(
        "Ownership stakes",
        ["Company", "Company number", "Control (verbatim PSC band)", "Current/Former", "Ceased", "Source"],
        [[s.company_name, s.company_number, "; ".join(s.controls),
          "Current", "", s.source_url] for s in sig.stakes]
        + [[s.company_name, s.company_number, "; ".join(s.controls),
            "Former", s.ceased_on, s.source_url] for s in sig.former_stakes],
        link_cols={6},
        note="No ownership stakes matched to this person.",
    )

    # --- Stake value estimates ---
    est = valuation.build_estimates(sheet, sheet.confirmed_name)
    est_rows = []
    for s in est.stakes:
        est_rows.append([
            s.company_name, s.company_number,
            f"{s.stake_lo:.0%}–{s.stake_hi:.0%}",
            s.net_assets if s.net_assets is not None else "not machine-readable",
            s.accounts_date,
            s.value_lo, s.value_hi, s.source_url,
        ])
    ws_est = _add_sheet(
        "Stake estimates",
        ["Company", "Number", "Stake band (PSC)", "Net assets (GBP)",
         "Accounts to", "Est. value low (GBP)", "Est. value high (GBP)", "Source"],
        est_rows,
        link_cols={8},
        note="No disclosed share-ownership stakes to value.",
    )
    if est_rows:
        r = len(est_rows) + 3
        note_cell = ws_est.cell(row=r, column=1)
        note_cell.value = (
            "ESTIMATES, not facts. Method: stake band (Companies House PSC) × net "
            "assets from the latest filed accounts (book value — a conservative "
            "floor, not market value). "
            f"Total where computable: £{est.total_lo:,.0f} – £{est.total_hi:,.0f}."
        )
        note_cell.font = Font(italic=True, color="6B7280")

    # --- Companies ---
    _add_sheet(
        "Companies",
        ["Company", "Number", "Status", "Incorporated", "Accounts to",
         "Revenue (GBP)", "Pre-tax profit (GBP)", "Employees", "Net assets (GBP)",
         "Has charges", "Insolvency history", "Source"],
        [[c.company_name, c.company_number, c.status,
          c.incorporation_date, c.accounts_last_made_up_to,
          c.turnover, c.profit_before_tax, c.employees, c.net_assets,
          c.has_charges, c.has_insolvency_history, c.source_url]
         for c in sheet.companies],
        link_cols={12},
        note="No company profiles returned.",
    )

    # --- Political donations ---
    _add_sheet(
        "Donations",
        ["Date", "Value (GBP)", "Recipient", "Donor name on record", "Source"],
        [[d.date, d.value, d.recipient, d.donor_name, d.source_url]
         for d in sheet.donations],
        link_cols={5},
        note="No registered political donations matched (Electoral Commission).",
    )

    # --- PSC (all, incl. others) ---
    _add_sheet(
        "PSC (all)",
        ["Company", "PSC name", "Kind", "Natures of control", "Notified", "Ceased", "Source"],
        [[p.company_name, p.name, p.kind, "; ".join(p.natures_of_control),
          p.notified_on, p.ceased_on, p.source_url] for p in sheet.psc_filings],
        link_cols={7},
        note="No PSC filings returned.",
    )

    # --- Regulatory (FCA) ---
    _add_sheet(
        "Regulatory (FCA)",
        ["Name", "IRN", "Status", "Roles", "Firms", "Source"],
        [[r.name, r.reference_number, r.status, ", ".join(r.roles),
          ", ".join(r.firms), r.source_url] for r in sheet.fca_records],
        link_cols={6},
        note="No FCA-approved individual matched (or FCA not configured).",
    )

    # --- Charities ---
    _add_sheet(
        "Charities",
        ["Name", "Charity number", "Status", "Source"],
        [[c.name, c.charity_number, c.status, c.source_url] for c in sheet.charities],
        link_cols={4},
        note="No matching charities (or Charity Commission not configured).",
    )

    # --- Sanctions / PEP ---
    _add_sheet(
        "Sanctions & PEP",
        ["Name", "Type", "Topics", "Countries", "Score", "Source"],
        [[h.name, h.schema, ", ".join(h.topics), ", ".join(h.countries),
          h.score, h.source_url] for h in sheet.sanctions_hits],
        link_cols={6},
        note="No sanctions/PEP matches (or OpenSanctions not configured).",
    )

    # --- Gazette ---
    _add_sheet(
        "Gazette notices",
        ["Title", "Published", "Link"],
        [[g.title, (g.published or "")[:10], g.link] for g in sheet.gazette_notices],
        link_cols={3},
        note="No Gazette notices matched.",
    )

    # --- News ---
    _add_sheet(
        "News",
        ["Title", "Source", "Published", "Link"],
        [[n.title, n.source, n.published, n.link] for n in sheet.news],
        link_cols={4},
        note="No news results.",
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
